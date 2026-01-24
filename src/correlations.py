import argparse
import traceback # for pretty-printingof  any issues that happened during runtime; if we hit FileNotFound I don't appreciate when a log traceback is shown, the error should be simple and clear
import sys # for printing out error messages if something is going wrong

from datetime import datetime # for logging
import time # for performance measures
from pathlib import Path
import re

import pandas as pd
import numpy as np



err_import_pyreadstat = None
try:
    import pyreadstat
except ImportError as err:
    err_import_pyreadstat = err


from scipy.stats import chi2_contingency



err_import_mdd = None
try:
    import win32com.client
except ImportError as err:
    err_import_mdd = err


from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # for styling resulting excel
from openpyxl.formatting.rule import FormulaRule, CellIsRule # for styling resulting excel



MIN_CASES_LIMIT = 8 # hard-coded, for now





def phi_coefficient(x, y):
    """Compute Phi for two binary Series."""
    table = pd.crosstab(x, y)
    if table.shape != (2,2):
        # Pad table if a value is missing
        table = table.reindex(index=[0,1], columns=[0,1], fill_value=0)
    chi2, p, _, _ = chi2_contingency(table)
    phi = np.sqrt(chi2 / table.to_numpy().sum())
    return phi




def read_file_pyreadstat(input_filename,group_filter):
    input_filename = Path.resolve(input_filename)
    if not(Path(input_filename).is_file()):
        raise FileNotFoundError('file not found: {fname}'.format(fname=input_filename))
    input_filename_str = '{f}'.format(f=input_filename)
    if err_import_pyreadstat:
        raise err_import_pyreadstat
    print('reading file with pyreadstat...')
    df, meta = pyreadstat.read_sav(input_filename_str)
    if group_filter:
        print('query data with filter {f}...'.format(f=group_filter))
        df = df.query(group_filter)
    return df, meta

def read_file_mdd(input_filename,group_filter):
    def sanitize_name_for_sql_string(n):
        return '{n}'.format(n=n).replace('"','')
    # def should_process_column(col):
    #     return True
    # def detect_col_type(col):
    #     return 'numeric'
    def iter_safe(series): # fucking pythonic ugly hack; as far as I understand, there is no better workaround
        try:
            for v in series:
                yield v
        except (AttributeError, TypeError) as e:
            msg = str(e)
            if 'total_seconds' in msg or 'ints_to_pydatetime' in msg:
                try:
                    # Fallback: bypass pandas datetime machinery
                    # vals = series.array.to_numpy(dtype=object, copy=False)
                    # if pd.api.types.is_datetime64tz_dtype(series):
                    if isinstance(series.dtype,pd.DatetimeTZDtype):
                        series = series.dt.tz_localize(None)
                    vals = series._values
                    for v in vals:
                        yield v
                except:
                    vals = pd.Series([pd.NaT]*len(series))
                    for v in vals:
                        yield v
            else:
                raise
    class MDDDocument:
        """Class that wraps connection to MDD"""
        class Error(Exception):
            """For errors accessing MDD"""
            pass
        def __init__(self,mdd_path,method='open',config={}):

            self.__document = None

            if method=='open':
                # mDocument = win32com.client.Dispatch("MDM.Document")
                mDocument = win32com.client.Dispatch("MDM.Document")
                # openConstants_oNOSAVE = 3
                openConstants_oREAD = 1
                # openConstants_oREADWRITE = 2
                print('opening MDD document using method "open": "{path}"'.format(path=mdd_path))
                # we'll check that the file exists so that the error message is more informative - otherwise you see a long stack of messages that do not tell much
                if not(Path(mdd_path).is_file()):
                    raise FileNotFoundError('file not found: {fname}'.format(fname=mdd_path))
                mDocument.Open( '{path}'.format(path=Path(mdd_path).resolve()), '', openConstants_oREAD )
                self.__document = mDocument
            elif method=='join':
                # mDocument = win32com.client.Dispatch("MDM.Document")
                mDocument = win32com.client.Dispatch("MDM.Document")
                print('opening MDD document using method "join": "{path}"'.format(path=mdd_path))
                # we'll check that the file exists so that the error message is more informative - otherwise you see a long stack of messages that do not tell much
                if not(Path(mdd_path).is_file()):
                    raise FileNotFoundError('file not found: {fname}'.format(fname=mdd_path))
                mDocument.Join('{path}'.format(path=Path(mdd_path).resolve()), "{..}", 1, 32|16|512)
                self.__document = mDocument
            else:
                raise Exception('MDD Open: Unknown open method, {method}'.format(method=method))

            config_default = {
            }
            self.__config = {
                **config_default,
                **config,
                'mdd_path': mdd_path,
                'read_datetime': datetime.now(),
            }
            self.__category_map_cache = {}
        def __del__(self):
            try:
                if self.__document is not None:
                    self.__document.Close()
            except:
                pass
            print('MDD document closed')
        def __enter__(self):
            return self
        def __exit__(self, exc_type, exc_val, exc_tb):
            pass
        def code_to_category_name(self,code):
            code = int(code)
            if code in self.__category_map_cache:
                return self.__category_map_cache[code]
            result = self.__document.CategoryMap.ValueToName(code)
            self.__category_map_cache[code] = result
            return result
    def clean_data(df,cb_category_map):
        def mdd_cat_parse(resp):
            if not resp:
                return []
            resp = re.sub(r'^\s*\{\s*(.*?)\s*\}\s*$',lambda m: m[1],str(resp))
            resp = resp.split(',')
            resp = [int(str(v).strip()) for v in resp if v.strip()]
            return resp
        def normalize_key(v):
            if isinstance(v, (pd.Timestamp, np.datetime64)):
                return '{v}'.format(v=v)
            elif pd.isna(v):
                return ''
            else:
                try:
                    hash(v)
                    return v
                except:
                    return '{v}'.format(v=v)
        def map_category(cat_code):
            # return 'cat_{s}'.format(s=cat_code)
            return '{s}'.format(s=cb_category_map(cat_code))
        def clean_column_convert_categorical(series):
            result = []
            for val in iter_safe(series):
                codes = mdd_cat_parse(val)
                codes_mapped = [ map_category(c) for c in codes ]
                val_mapped = '{' + ', '.join(codes_mapped) + '}'
                result.append(val_mapped)
            return pd.Series(result)
        def clean_column(series):
            mdd_raw_pattern = re.compile(r'^\s*\{\s*(?:\d+(?:\s*,\s*\d+)*)?\s*\}\s*$')
            all_values = { normalize_key(v): True for v in iter_safe(series) }
            col_type = None
            values_non_blank = [v for v in all_values.keys() if not not (str(v).strip())]
            if values_non_blank and all(mdd_raw_pattern.match(str(v)) for v in values_non_blank):
                col_type = 'categorical'
            if col_type == 'categorical':
                return clean_column_convert_categorical(series)
            else:
                return series
        print('converting columns to categorical type with real category names...')
        performance_counter = iter(PerformanceMonitor(config={
            'total_records': len(df.columns),
            'report_frequency_records_count': 20,
            'report_frequency_timeinterval': 6,
            'report_text_pipein': 'convertig variables',
        }))
        # TODO: why are we converting all, when we don't need all? To waste processing time?
        for col in df.columns:
            col_name = '{f}'.format(f=col)
            try:
                next(performance_counter)
                df[col] = clean_column(df[col])
            except Exception as e:
                print('MDD: trying to convert categories: failed on column "{c}"'.format(c=col_name,file=sys.stderr))
                raise e
        return df
    def read_data(input_filename,group_filter):
        input_filename = Path.resolve(input_filename)
        input_filename_mdd = Path(input_filename).with_suffix('.mdd').resolve()
        if not input_filename_mdd.is_file():
            raise FileNotFoundError('MDD File not found: {f}'.format(f=input_filename_mdd))
        input_filename_data = Path(input_filename).with_suffix('.ddf').resolve()
        if not input_filename_data.is_file():
            raise FileNotFoundError('DDF File not found: {f}'.format(f=input_filename_data))
        if err_import_mdd:
            raise err_import_mdd
        sfilter = group_filter or 'true'
        try:
            print('loading with ADODB connection...')
            oconnection = win32com.client.Dispatch('ADODB.Connection')
            try:
                orecordset = win32com.client.Dispatch('ADODB.Recordset')
                oconnection.Open("Provider=mrOleDB.Provider.2;Data Source=mrDataFileDsc;Location="+sanitize_name_for_sql_string(input_filename_data)+";Initial Catalog="+sanitize_name_for_sql_string(input_filename_mdd)+";MR Init MDM Access=0;")
                orecordset.Open("SELECT * FROM VDATA WHERE " + sanitize_name_for_sql_string(sfilter) , oconnection)
                cols = [orecordset.Fields(i).Name for i in range(orecordset.Fields.Count)]
                data = orecordset.GetRows()
                df = pd.DataFrame(list(zip(*data)),columns=cols)
                # if group_filter:
                #     print('query data with filter {f}...'.format(f=group_filter))
                #     df = df.query(group_filter)
                return df, None # "finally" will run anyway
            finally:
                try:
                    if orecordset is not None:
                        orecordset.Close()
                    orecordset = None
                except:
                    pass
        finally:
            try:
                if oconnection is not None:
                    oconnection.Close()
                oconnection = None
                print('ADODB connection closed.')
            except:
                pass
    print('reading MDD/DDF...')
    # return clean_data(read_data(input_filename,group_filter))
    mdmdoc = MDDDocument(Path.resolve(input_filename).with_suffix('.mdd'))
    df, meta = read_data(input_filename,group_filter)
    df = clean_data(df,cb_category_map=lambda code: mdmdoc.code_to_category_name(code)) # mdmdoc.CategoryMap.ValueToName(int(2011))
    return df, meta

def read_file_oledb(input_filename,group_filter):
    input_filename = Path.resolve(input_filename)
    if not(Path(input_filename).is_file()):
        raise FileNotFoundError('file not found: {fname}'.format(fname=input_filename))
    print('reading file with OLDEBD connection...')
    raise Exception('This reading method is not implemented yet')

def read_file_notimplemented_placeholder_future(input_filename,group_filter):
    input_filename = Path.resolve(input_filename)
    if not(Path(input_filename).is_file()):
        raise FileNotFoundError('file not found: {fname}'.format(fname=input_filename))
    raise Exception('This reading method is not implemented yet')


readers = {
    'pyreadstat': read_file_pyreadstat,
    'readermdd': read_file_mdd,
    'oledb': read_file_oledb,
    'notimplemented': read_file_notimplemented_placeholder_future,
}


def read_file(input_filename,format=None,group_filter=None):

    reader = None
    input_filename = Path(input_filename)

    if format is None:
        format = 'autodetect'
    if format and format=='mdd':
        format = 'mdd'
    elif format and format=='spss':
        format = 'spss'
    elif format and format=='csv':
        format = 'csv'
    elif format and format=='autodetect':
        filetype = Path(input_filename).suffix.strip().lower()
        if filetype in ['.sav']:
            format = 'spss'
        elif filetype in ['.mdd','.ddf']:
            format = 'mdd'
        elif filetype in ['.csv','.tsv']:
            format = 'csv'
        else:
            raise Exception('auto-detect failed: input file format not recognized, or can\'t be handled yet: {fname}'.format(fname=Path(input_filename).name))

    reader = None
    if format=='spss':
        reader = 'pyreadstat'
    elif format=='mdd':
        reader = 'readermdd'
    elif format=='csv':
        reader = 'notimplemented'
    else:
        raise Exception('format not supported: {fmt}'.format(fmt=format))

    if reader in readers:
        return readers[reader](input_filename,group_filter)
    else:
        raise Exception('Reader not found: {f}'.format(f=reader))


def save_results(results,out_filename,config):
    def format_sheet_overview(sheet):
        font_main_transparent = Font(color='FFFFFF')
        fill_main_transparent = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        # font_header = Font(color='013E40')
        # border_main = Border(left=Side(style='thick',color='ffffff'),right=Side(style='thick',color='ffffff'),bottom=Side(style='thick',color='ffffff'),top=Side(style='thick',color='ffffff'),)
        border_noborder = Border()
        # # alignment_center = Alignment(horizontal='center',vertical='top')
        # # alignment_indent = Alignment(indent=2,vertical='top')
        alignment_col1 = Alignment(indent=2,horizontal='right',vertical='top')
        alignment_col2 = Alignment(indent=2,horizontal='left',vertical='top')
        font_header = Font(color="666666",bold=True,size=20)
        for row_number, row in enumerate([r for r in sheet.rows][1:]):
            sheet.row_dimensions[row_number+2].height = 25
            for cell in row:
                # cell.fill = fill_main
                # cell.font = font_main
                cell.border = border_noborder
                row[0].alignment = alignment_col1
                row[1].alignment = alignment_col2
        for row in sheet.iter_rows(min_row=1,max_row=1,min_col=1,max_col=26):
            for cell in row:
                cell.fill = fill_main_transparent # fill_header
                cell.font = font_main_transparent # font_header
                cell.border = border_noborder
        # for row in sheet.iter_rows(min_row=2,max_row=999,min_col=1,max_col=26):
        #     for cell in row:
        #         # cell.fill = fill_main
        #         # cell.font = font_main
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 125    
        sheet.row_dimensions[1].height = 20
        sheet.row_dimensions[2].height = 35
        sheet['A1'].fill = fill_main_transparent
        sheet['B1'].fill = fill_main_transparent
        sheet['A1'].font = font_main_transparent
        sheet['B1'].font = font_main_transparent
        sheet['A2'].border = border_noborder
        sheet['B2'].border = border_noborder
        sheet["B2"].font = font_header
    def format_sheet_phi_coefficients(sheet):
        alignment_left_top = Alignment(indent=0,horizontal='left',vertical='top')
        alignment_left_bottom = Alignment(indent=0,horizontal='left',vertical='bottom')
        fill_strong = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
        fill_light = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
        for row in sheet.iter_rows(min_col=1, max_col=1, min_row=1, max_row=sheet.max_row):
            for cell in row:
                cell.font = Font(bold=False)
                cell.border = Border()
                cell.alignment = alignment_left_top
        for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
            for cell in col:
                cell.font = Font(bold=True)
                cell.border = Border()
                cell.alignment = alignment_left_bottom
        for col in sheet.columns:
            col_letter = col[0].column_letter  # Get the column letter (e.g., 'A', 'B', etc.)
            # sheet.column_dimensions[col_letter].width = 80
        min_row = 2
        min_col = 2
        max_row = sheet.max_row
        max_col = sheet.max_column
        if max_row<min_row:
            max_row = min_row
        if max_col<min_col:
            max_col=min_col
        if (max_row >= min_row) and (max_col >= min_col):
            start_cell = sheet.cell(row=min_row, column=min_col).coordinate
            end_cell = sheet.cell(row=max_row, column=max_col).coordinate
            cell_range = '{start_cell}:{end_cell}'.format(start_cell=start_cell,end_cell=end_cell)
            rule_strong = CellIsRule(operator="greaterThan", formula=["0.5"], fill=fill_strong)
            rule_moderate = CellIsRule(operator="greaterThan", formula=["0.3"], fill=fill_light)
            sheet.conditional_formatting.add(cell_range, rule_strong)
            sheet.conditional_formatting.add(cell_range, rule_moderate)
        sheet.column_dimensions['A'].width = 25
        sheet.row_dimensions[1].height = 21
    def format_sheet_strong_results(sheet):
        alignment_left_top = Alignment(indent=0,horizontal='left',vertical='top')
        alignment_left_bottom = Alignment(indent=0,horizontal='left',vertical='bottom')
        for row in sheet.iter_rows(min_col=1, max_col=1, min_row=1, max_row=sheet.max_row):
            for cell in row:
                cell.font = Font(bold=False)
                cell.border = Border()
                cell.alignment = alignment_left_top
        for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
            for cell in col:
                cell.font = Font(bold=True)
                cell.border = Border()
                cell.alignment = alignment_left_bottom
        for col in sheet.columns:
            col_letter = col[0].column_letter  # Get the column letter (e.g., 'A', 'B', etc.)
            # sheet.column_dimensions[col_letter].width = 80
        sheet.column_dimensions['A'].width = 35
        sheet.row_dimensions[1].height = 21
    df_phi_matrix = results['df_phi_matrix']
    df_strong_results = results['df_strong_results']
    with pd.ExcelWriter(out_filename, engine='openpyxl') as writer:
        # overview sheet
        df_overview = pd.DataFrame([
            ['',Path(config['input_filename']).name],
            ['Input File:',config['input_filename']],
            ['Run Datetime:','{dt}'.format(dt=config['time_start'])],
            ['Case Data Filter:',config['group_filter']],
            ['Records within Case Data Filter',config['cases_count']],
            ['Variable Select Pattern:',config['var_pattern']],
            ['Variables Analyzed:',config['variables_analyzed']],
        ], columns=['name','value']).set_index("name")
        df_overview.to_excel(writer, sheet_name='overview')
        format_sheet_overview(writer.sheets['overview'])
        # phi_coefficients sheet
        df_phi_matrix.to_excel(writer, sheet_name='phi_coefficients')
        format_sheet_phi_coefficients(writer.sheets['phi_coefficients'])
        # df_strong_results sheet
        df_results = pd.DataFrame([[f"{i} vs {j}",phi,count] for i,j,phi,count in df_strong_results], columns=['pair','phi','cases']).set_index("pair")
        df_results.to_excel(writer, sheet_name='df_strong_results')
        format_sheet_strong_results(writer.sheets['df_strong_results'])




def create_DVs(df):
    """ nothing here at the moment but can be potentially expanded"""
    # print('generating derived variables...')
    # # this is abviously generated by chatgpt but this is good, no need to feel sorry
    # m_to_dvqo = {
    #     'M5': '001',
    #     'M6': '002',
    #     'M7': '003',
    #     'M8': '004',
    #     'M9': '005'
    # }
    # # Loop through all columns starting with M5-M9
    # for col in df.columns:
    #     # Match columns like M5_001_001, M6_042_017, etc.
    #     if re.match(VAR_PATTERN, col) and not re.match(VAR_EXCLUDE_PATTERN, col):
    #         # Split parts: M-question, attribute, brand
    #         parts = col.split('_')
    #         m_question = parts[0]        # 'M5'
    #         attribute = parts[1]         # '001'..'015', etc.
    #         brand = parts[2]             # '001'..'005', etc.
    #         # Get corresponding DVQO question number
    #         dvqo_qnum = m_to_dvqo[m_question]
    #         # Build DVQO column name
    #         dvqo_col = f'DV_QuestionOrder_{dvqo_qnum}_{brand}'
    #         # Build new DVM column name
    #         dvm_col = f'DV{m_question}_{attribute}_{brand}'
    #         # Create derived column conditionally
    #         if dvqo_col in df.columns:
    #             df[dvm_col] = df[col].where(df[dvqo_col] == 1, other=np.nan)
    return df



class PerformanceMonitor:
    def __init__(self,config={}):
        self.__config = config
        self.progress = None
        self.time_started = None
        self.time_last_reported = None
        self.progress_last_reported = None
        self.userprovideddata_totalrecords = config['total_records'] if 'total_records' in config else None
        self.config_frequency_records = config['report_frequency_records_count'] if 'report_frequency_records_count' in config else None
        self.config_frequency_timeinterval = config['report_frequency_timeinterval'] if 'report_frequency_timeinterval' in config else None
        self.config_text_pipein = config['report_text_pipein'] if 'report_text_pipein' in config and config['report_text_pipein'] else 'progress'
    def __iter__(self):
        self.progress = 0
        self.time_started = time.time()
        self.time_last_reported = self.time_started
        self.progress_last_reported = -1
        return self
    def _calc_eta(self,time_now=None):
        def calc_eta(time_start,time_now,records_expected_total,records_now):
            return (1*time_start+int((time_now-time_start)*(records_expected_total/records_now)))
        if self.userprovideddata_totalrecords is None:
            return None
        if not time_now:
            time_now = time.time()
        time_started = self.time_started
        return calc_eta(time_started,time_now,self.userprovideddata_totalrecords,self.progress)
    def __next__(self):
        def fmt_duration(seconds):
            def fmt(v):
                v = '{v}'.format(v=v)
                return re.sub(r'(\.[1-9]\d*?)[0]{3}\d*',lambda m: m[1],v,flags=re.I|re.DOTALL)
            if seconds<300:
                return '{n} seconds'.format(n=fmt(int(seconds)))
            else:
                if seconds<6000:
                    return '{n} minutes'.format(n=fmt(0.1*int(seconds/6)))
                else:
                    return '{n} hours'.format(n=fmt(0.1*int(seconds/360)))
        self.progress = self.progress + 1
        if (self.config_frequency_records is None) or (self.progress - self.progress_last_reported > self.config_frequency_records):
            time_now = time.time()
            if (self.config_frequency_timeinterval is None) or ((time_now - self.time_last_reported)>self.config_frequency_timeinterval):
                eta = self._calc_eta(time_now)
                print( '{text_pipe}: processing {nline}{display_out_total}{display_details}'.format(
                    nline = self.progress,
                    display_out_total = ' / {nlinetotal}'.format(nlinetotal=self.userprovideddata_totalrecords) if (self.userprovideddata_totalrecords is not None) else '',
                    display_details = ' ({per}%{details_eta})'.format(
                        per = round(self.progress*100/self.userprovideddata_totalrecords,1),
                        details_eta = ', remaining: {remaining}, ETA: {eta}'.format(
                            remaining = fmt_duration(eta-time_now),
                            eta = '{t}'.format(t=time.strftime('%m/%d/%Y %H:%M:%S',time.localtime(eta))),
                        ) if eta else '',
                    ) if (self.userprovideddata_totalrecords is not None) else '',
                    text_pipe = self.config_text_pipein,
                ))
                self.progress_last_reported = self.progress
                self.time_last_reported = time_now
        return None




def clean_column_in_dataframe(series):
    def normalize_key(v):
        if isinstance(v, (pd.Timestamp, np.datetime64)):
            return '{v}'.format(v=v)
        elif pd.isna(v):
            return ''
        else:
            try:
                hash(v)
                return v
            except:
                return '{v}'.format(v=v)
    def iter_safe(series): # fucking pythonic ugly hack; as far as I understand, there is no better workaround
        try:
            for v in series:
                yield v
        except (AttributeError, TypeError) as e:
            msg = str(e)
            if 'total_seconds' in msg or 'ints_to_pydatetime' in msg:
                try:
                    # Fallback: bypass pandas datetime machinery
                    # vals = series.array.to_numpy(dtype=object, copy=False)
                    # if pd.api.types.is_datetime64tz_dtype(series):
                    if isinstance(series.dtype,pd.DatetimeTZDtype):
                        series = series.dt.tz_localize(None)
                    vals = series._values
                    for v in vals:
                        yield v
                except:
                    vals = pd.Series([pd.NaT]*len(series))
                    for v in vals:
                        yield v
            else:
                raise
    result = pd.DataFrame()
    # count = len(series)
    col_type = None
    all_values = { normalize_key(v): True for v in iter_safe(series) }
    mdd_raw_pattern = re.compile(r'^\s*\{\s*(?:\d+(?:\s*,\s*\d+)*)?\s*\}\s*$')
    mdd_convertedcatnames_pattern = re.compile(r'^\s*\{\s*(?:\w+(?:\s*,\s*\w+)*)?\s*\}\s*$')
    if (1 in all_values) and len(set(all_values.keys())-{0,1})==0:
        col_type = 'multipunch_flag'
    elif len(set(all_values.keys())-{0,1})==0:
        col_type = 'blank'
    elif all(not (str(v).strip()) or mdd_raw_pattern.match(str(v)) for v in all_values.keys()):
        col_type = 'categorical_mdd'
    elif all(not (str(v).strip()) or mdd_convertedcatnames_pattern.match(str(v)) for v in all_values.keys()):
        col_type = 'categorical_withconvertedcatnames_mdd'
    else:
        col_type = 'need_categorize'
    if col_type == 'multipunch_flag':
        result['@'] = series.fillna(0).astype(int)
    elif col_type == 'blank':
        result['@'] = series.fillna(0).astype(int)
    elif col_type == 'need_categorize':
        code_index = 0
        for code, _ in all_values.items():
            values = [ 1 if normalize_key(v)==code else 0 for v in iter_safe(series)]
            # col_name_created = '@_{n}'.format(n=str(code_index).zfill(3))
            col_name_created = '@ =* {n}'.format(n=code)
            result[col_name_created] = pd.Series(values)
            code_index += 1
    elif col_type=='categorical_mdd':
        def mdd_cat_parse(resp):
            resp = re.sub(r'^\s*\{\s*(.*?)\s*\}\s*$',lambda m: m[1],resp)
            resp = resp.split(',')
            resp = [int(str(v).strip()) for v in resp if v.strip()]
            return resp
        all_values = { normalize_key(v): True for resp in iter_safe(series) for v in mdd_cat_parse(resp) }
        code_index = 0
        for code, _ in all_values.items():
            values = [ 1 if code in mdd_cat_parse(v) else 0 for v in iter_safe(series)]
            # col_name_created = '@_{n}'.format(n=str(code_index).zfill(3))
            col_name_created = '@ =* {{{n}}}'.format(n=code)
            result[col_name_created] = pd.Series(values)
            code_index += 1
    elif col_type=='categorical_withconvertedcatnames_mdd':
        def mdd_cat_parse(resp):
            resp = re.sub(r'^\s*\{\s*(.*?)\s*\}\s*$',lambda m: m[1],resp)
            resp = resp.split(',')
            resp = [str(v).strip() for v in resp if v.strip()]
            return resp
        all_values = { normalize_key(v): True for resp in iter_safe(series) for v in mdd_cat_parse(resp) }
        code_index = 0
        for code, _ in all_values.items():
            values = [ 1 if code in mdd_cat_parse(v) else 0 for v in iter_safe(series)]
            # col_name_created = '@_{n}'.format(n=str(code_index).zfill(3))
            col_name_created = '@ =* {{{n}}}'.format(n=code)
            result[col_name_created] = pd.Series(values)
            code_index += 1
    else:
        raise Exception('parsing values as "{fmtaa}": not implemented, or format not recognized'.format(fmt=col_type))

    return result

def prepare_df(df,pattern_check,config):
    print('preparing df (selecting variables)...')
    stub_cols = [col for col in df.columns if pattern_check(col)]
    print('FYI variables we take for analysis: [ {vars} ]'.format(vars=', '.join(['{c}'.format(c=c) for c in stub_cols])))
    df = df[stub_cols].copy()

    # Ensure values are 0/1
    processed_data = pd.DataFrame()
    for col in stub_cols:
        col_name = '{c}'.format(c=col)
        try:
            series = df[col]
            # try:
            result_df = clean_column_in_dataframe(series)
            for col_append,col_appended in zip(result_df.columns.str.replace('@',col_name),result_df.columns):
                processed_data[col_append] = result_df[col_appended]
            # except Exception as e:
            #     raise Exception('ERROR: Column "{c}": df should follow the format with values 0/1 (for multi-punch), or numeric. Otherwise we can\'t check for correlations. Failed on column: "{c}". The original error message: "{msg}"'.format(msg=e,c=col_name)) from e
        except Exception as e:
            print('ERROR: Trying to clean column "{c}" and convert to 1/0 number and failed'.format(c=col_name),file=sys.stderr)
            raise e
    df = processed_data
    return df


def compute(df,config):

    stub_cols = [col for col in df.columns]

    phi_matrix = pd.DataFrame(index=stub_cols, columns=stub_cols, dtype=float)
    strong_results = []

    print('using chi2_contingency()...')
    performance_counter = iter(PerformanceMonitor(config={
        'total_records': len(stub_cols)**2,
        'report_frequency_records_count': 20,
        'report_frequency_timeinterval': 6
    }))
    for col1 in stub_cols:
        for col2 in stub_cols:
            try:
                next(performance_counter)
                if (col1 == col2):
                    phi_matrix.loc[col1, col2] = 1.0
                elif df[col1].nunique() > 1 and df[col2].nunique() > 1 and df[col1].count()>MIN_CASES_LIMIT and df[col2].count()>MIN_CASES_LIMIT:
                    if pd.isna(phi_matrix.loc[col1, col2]):
                        phi = phi_coefficient(df[col1], df[col2])
                        phi_matrix.loc[col1, col2] = phi
                        phi_matrix.loc[col2, col1] = phi  # symmetric
                        strong_results.append([col1,col2,phi,len(df[[col1, col2]].dropna())])
                    else:
                        # should have been already populated
                        pass
                else:
                    # skip
                    phi_matrix.loc[col1, col2] = None
            except Exception as e:
                print('Error: failed when processing `{f1}` and `{f2}` variables'.format(f1=col1,f2=col2),file=sys.stderr)
                raise e
    
    # sort
    is_identical = lambda col1,col2,*args:(col1 == col2)
    strong_results = [r for r in strong_results if not (is_identical(*r))]

    strong_results = sorted(strong_results,key=lambda a:-a[2])
    PHI_THRESHOLD = 0.01
    PHI_THRESHOLD_LIMIT = len(stub_cols)**2+999 # this was designed to limit output to first n records, like when it is printed in command line... But we don't have to limit, we can print all. So I set the limit to something higher than theoretically possible
    if len(strong_results)>PHI_THRESHOLD_LIMIT:
        if strong_results[PHI_THRESHOLD_LIMIT][2]>PHI_THRESHOLD:
            strong_results = [r for r in strong_results if r[2]>PHI_THRESHOLD]
        else:
            strong_results = strong_results[:PHI_THRESHOLD_LIMIT]
    
    return phi_matrix, strong_results




def main():
    try:
        time_start = datetime.now()
        script_name = 'test correlations'

        parser = argparse.ArgumentParser(
            description='A program to check correlations in spss data'
        )
        parser.add_argument('--inpfile', required=True, help='Input SPSS data file')
        parser.add_argument('--outfile', required=False, help='Output Excel file with results')
        parser_pattern_group = parser.add_mutually_exclusive_group(required=True)
        parser_pattern_group.add_argument('--pattern_regex', help='Regex pattern to select questions; like "^M(?:5|6|7|8|9)_\d{3}_007"')
        parser_pattern_group.add_argument('--pattern_mask', help='Wildcard pattern to select questions; like "^M*_007"')
        parser.add_argument('--filter', required=False, help='Filter to filter down sample; for example, "DV_LinkType != 2" or "DV_LinkType == 2"')
        parser.add_argument(
            '--format',
            default='autodetect',
            help='Data file type, or "autodetect"',
            choices=['autodetect','mdd','spss','csv'],
            type=str,
            required=False
        )

        args = parser.parse_args()
        config = {}

        print('{script_name}: script started at {dt}'.format(dt=time_start,script_name=script_name))
        config['time_start'] = time_start
        config['script_name'] = script_name



        # =============================
        # 1. Read SPSS file
        # =============================
        input_filename = None
        if args.inpfile:
            input_filename = Path(args.inpfile)
            input_filename = input_filename.resolve()
        format = None
        if args.format:
            format = args.format
        config['input_filename'] = input_filename

        group_filter = None
        if args.filter:
            group_filter = args.filter
        config['group_filter'] = group_filter

        print('loading input file {f}...'.format(f=input_filename))
        df, meta = read_file('{f}'.format(f=input_filename),format,group_filter)
        config['df_meta'] = meta
        
        
        
        # =============================
        # 2.1.Create DVs
        # =============================
        print('creating derived variables, if needed...')
        df = create_DVs(df)

        # # =============================
        # # 2.2. Apply group filter (optional)
        # # =============================
        # group_filter = None
        # if args.filter:
        #     group_filter = args.filter
        # if group_filter:
        #     print('query data with filter {f}...'.format(f=group_filter))
        #     if group_filter:
        #         df = df.query(group_filter)
        # config['group_filter'] = group_filter
        # config['cases_count'] = len(df.index)

        # =============================
        # 2.3. Select and prepare variables
        # =============================
        var_pattern = None
        pattern_check = lambda col: True
        if args.pattern_regex:
            var_pattern = args.pattern_regex
            pattern_check = lambda col: re.match(var_pattern, col,flags=re.I)
        if args.pattern_mask:
            var_pattern = args.pattern_mask
            if not re.match(r'^[\w\*\?]+$',var_pattern,flags=re.I):
                raise Exception('Pattern mask format: only alphanumeric characters, "?" and "*" are allowed: "{p}". If you need more flexibility please kindly use --pattern_regex'.format(p=var_pattern))
            pattern_check = lambda col: re.match(r'^'+re.sub(r'\*','.*',re.sub(r'\?','.',var_pattern))+r'$', col,flags=re.I)
        config['var_pattern'] = var_pattern

        print('computing...')
        config['cases_count'] = len(df.index)
        print('FYI: records in case data within filter: {n}'.format(n=config['cases_count']))
        df = prepare_df(df,pattern_check,config)
        stub_cols = [col for col in df.columns]
        variables_analyzed = '[ {vars} ]'.format(vars=', '.join(['{c}'.format(c=c) for c in stub_cols]))
        config['variables_analyzed'] = variables_analyzed



        # =============================
        # 3. Compute Phi correlation matrix
        # =============================
        config['df'] = df
        phi_matrix, strong_results = compute(df,config)
        results = {
            'df_phi_matrix': phi_matrix,
            'df_strong_results': strong_results,
        }



        # =============================
        # 4. Save
        # =============================
        out_filename = None
        if args.outfile:
            out_filename = Path(args.outfile)
        else:
            out_filename = Path('{f}_correlations.xlsx'.format(f=re.sub(r'\.(?:sav|spss)\s*?$','','{f}'.format(f=input_filename))))
        out_filename = Path.resolve(out_filename)
        config['out_filename'] = out_filename

        print('saving results as {f}...'.format(f=out_filename))
        save_results(
            results,
            '{f}'.format(f=out_filename),
            config,
        )




        time_finish = datetime.now()
        print('{script_name}: finished at {dt} (elapsed {duration})'.format(dt=time_finish,duration=time_finish-time_start,script_name=script_name))

    except Exception as e:
        # the program is designed to be user-friendly
        # that's why we reformat error messages a little bit
        # stack trace is still printed (I even made it longer to 20 steps!)
        # but the error message itself is separated and printed as the last message again

        # for example, I don't write 'print('File Not Found!');exit(1);', I just write 'raise FileNotFoundErro()'
        print('',file=sys.stderr)
        print('Stack trace:',file=sys.stderr)
        print('',file=sys.stderr)
        traceback.print_exception(e,limit=20)
        print('',file=sys.stderr)
        print('',file=sys.stderr)
        print('',file=sys.stderr)
        print('Error:',file=sys.stderr)
        print('',file=sys.stderr)
        print('{e}'.format(e=e),file=sys.stderr)
        print('',file=sys.stderr)
        exit(1)

if __name__ == '__main__':
    main()