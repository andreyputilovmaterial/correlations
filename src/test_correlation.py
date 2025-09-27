import argparse
import traceback

from datetime import datetime
from pathlib import Path
import re

import pandas as pd
import pyreadstat
import numpy as np
from scipy.stats import chi2_contingency

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule



MIN_CASES_LIMIT = 8 # hard-coded, for now




def phi_coefficient(x, y):
    """Compute Phi for two binary Series."""
    table = pd.crosstab(x, y)
    if table.shape != (2,2):
        # Pad table if a value is missing
        table = table.reindex(index=[0,1], columns=[0,1], fill_value=0)
    chi2, p, _, _ = chi2_contingency(table)
    phi = np.sqrt(chi2 / table.to_numpy().sum())
    return phi




def create_DVs(df):
    """ nothing here at the moment but can be potentially expanded"""
    # print('generating derived variables...')
    # m_to_dvqo = {
    #     'M5': '001',
    #     'M6': '002',
    #     'M7': '003',
    #     'M8': '004',
    #     'M9': '005'
    # }
    # # Loop through all columns starting with M5-M9
    # for col in df.columns:
    #     # Match columns like M5_001_001, M6_042_017, etc.
    #     if re.match(VAR_PATTERN, col) and not re.match(VAR_EXCLUDE_PATTERN, col):
    #         # Split parts: M-question, attribute, brand
    #         parts = col.split('_')
    #         m_question = parts[0]        # 'M5'
    #         attribute = parts[1]         # '001'..'015', etc.
    #         brand = parts[2]             # '001'..'005', etc.
    #         # Get corresponding DVQO question number
    #         dvqo_qnum = m_to_dvqo[m_question]
    #         # Build DVQO column name
    #         dvqo_col = f'DV_QuestionOrder_{dvqo_qnum}_{brand}'
    #         # Build new DVM column name
    #         dvm_col = f'DV{m_question}_{attribute}_{brand}'
    #         # Create derived column conditionally
    #         if dvqo_col in df.columns:
    #             df[dvm_col] = df[col].where(df[dvqo_col] == 1, other=np.nan)
    return df



def compute(df,var_pattern):
    print('prep data (select variables)...')
    stub_cols = [col for col in df.columns if re.match(var_pattern, col)]
    data = df[stub_cols].copy()

    # Ensure values are 0/1
    data = data.fillna(0).astype(int)

    phi_matrix = pd.DataFrame(index=stub_cols, columns=stub_cols, dtype=float)
    strong_results = []

    print('using chi2_contingency()...')
    for col1 in stub_cols:
        for col2 in stub_cols:
            if (col1 == col2):
                phi_matrix.loc[col1, col2] = 1.0
            elif data[col1].nunique() > 1 and data[col2].nunique() > 1 and data[col1].count()>MIN_CASES_LIMIT and data[col2].count()>MIN_CASES_LIMIT:
                if pd.isna(phi_matrix.loc[col1, col2]):
                    phi = phi_coefficient(data[col1], data[col2])
                    phi_matrix.loc[col1, col2] = phi
                    phi_matrix.loc[col2, col1] = phi  # symmetric
                    strong_results.append([col1,col2,phi,len(df[[col1, col2]].dropna())])
                else:
                    # should have been already populated
                    pass
            else:
                # skip
                phi_matrix.loc[col1, col2] = None
    
    # sort
    is_identical = lambda col1,col2,*args:(col1 == col2)
    strong_results = [r for r in strong_results if not (is_identical(*r))]

    strong_results = sorted(strong_results,key=lambda a:-a[2])
    PHI_THRESHOLD = 0.01
    PHI_THRESHOLD_LIMIT = len(stub_cols)**2+999 # this was designed to limit output to first n records, like when it is printed in command line... But we don't have to limit, we can print all. So I set the limit to something higher than theoretically possible
    if len(strong_results)>PHI_THRESHOLD_LIMIT:
        if strong_results[PHI_THRESHOLD_LIMIT][2]>PHI_THRESHOLD:
            strong_results = [r for r in strong_results if r[2]>PHI_THRESHOLD]
        else:
            strong_results = strong_results[:PHI_THRESHOLD_LIMIT]
    
    return phi_matrix, strong_results




def main():
    try:
        time_start = datetime.now()
        script_name = 'test correlations'

        parser = argparse.ArgumentParser(
            description='A program to check correlations in spss data'
        )
        parser.add_argument('--inpfile', required=True, help='Input SPSS data file')
        parser.add_argument('--outfile', required=False, help='Output Excel file with results')
        parser.add_argument('--pattern', required=True, help='Regex pattern to select questions; like "^M(?:5|6|7|8|9)_\d{3}_\d{3}"')
        parser.add_argument('--filter', required=False, help='Filter to filter down sample; for example, "DV_LinkType != 2" or "DV_LinkType == 2"')

        args = parser.parse_args()

        # Compile regex from string param
        pattern = re.compile(args.pattern)

        # with open(args.input, 'r') as f_in, open(args.output, 'w') as f_out:
        #     for line in f_in:
        #         if pattern.search(line):
        #             f_out.write(line)


        print('{script_name}: script started at {dt}'.format(dt=time_start,script_name=script_name))



        # =============================
        # 1. Read SPSS file
        # =============================
        input_filename = None
        if args.inpfile:
            input_filename = Path(args.inpfile)
        input_filename = Path.resolve(input_filename)
        if not(Path(input_filename).is_file()):
            raise FileNotFoundError('file not found: {fname}'.format(fname=input_filename))
        print('loading input file {f}...'.format(f=input_filename))
        df, meta = pyreadstat.read_sav(input_filename)
        
        
        
        # =============================
        # 2. Create DVs
        # =============================
        df = create_DVs(df)



        # =============================
        # 3. Apply group filter (optional)
        # =============================
        group_filter = None
        if args.filter:
            group_filter = args.filter
        if group_filter:
            print('query data with filter {f}...'.format(f=group_filter))
            if group_filter:
                df = df.query(group_filter)



        # =============================
        # 4. Compute Phi correlation matrix
        # =============================
        var_pattern = None
        if args.pattern:
            var_pattern = args.pattern
        print('computing...')
        phi_matrix, strong_results = compute(df,var_pattern)



        # =============================
        # 5. Save
        # =============================
        out_filename = None
        if args.outfile:
            out_filename = Path(args.outfile)
        else:
            out_filename = Path('{f}_correlations.xlsx'.format(f=re.sub(r'\.(?:sav|spss)\s*?$','','{f}'.format(f=input_filename))))
        out_filename = Path.resolve(out_filename)
        print('saving results as {f}...'.format(f=out_filename))
        with pd.ExcelWriter(out_filename, engine='openpyxl') as writer:
            def sanitize_text_extract_filename(s):
                return re.sub(r'^.*[/\\](.*?)\s*?$',lambda m: m[1],'{sstr}'.format(sstr=s))
            def format_sheet_overview(sheet):
                font_main_transparent = Font(color='FFFFFF')
                fill_main_transparent = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                # font_header = Font(color='013E40')
                # border_main = Border(left=Side(style='thick',color='ffffff'),right=Side(style='thick',color='ffffff'),bottom=Side(style='thick',color='ffffff'),top=Side(style='thick',color='ffffff'),)
                border_noborder = Border()
                # # alignment_center = Alignment(horizontal='center',vertical='top')
                # # alignment_indent = Alignment(indent=2,vertical='top')
                alignment_col1 = Alignment(indent=2,horizontal='right',vertical='top')
                alignment_col2 = Alignment(indent=2,horizontal='left',vertical='top')
                font_header = Font(color="666666",bold=True,size=20)
                for row_number, row in enumerate([r for r in sheet.rows][1:]):
                    sheet.row_dimensions[row_number+2].height = 25
                    for cell in row:
                        # cell.fill = fill_main
                        # cell.font = font_main
                        cell.border = border_noborder
                        row[0].alignment = alignment_col1
                        row[1].alignment = alignment_col2
                for row in sheet.iter_rows(min_row=1,max_row=1,min_col=1,max_col=26):
                    for cell in row:
                        cell.fill = fill_main_transparent # fill_header
                        cell.font = font_main_transparent # font_header
                        cell.border = border_noborder
                # for row in sheet.iter_rows(min_row=2,max_row=999,min_col=1,max_col=26):
                #     for cell in row:
                #         # cell.fill = fill_main
                #         # cell.font = font_main
                sheet.column_dimensions['A'].width = 35
                sheet.column_dimensions['B'].width = 125    
                sheet.row_dimensions[1].height = 20
                sheet.row_dimensions[2].height = 35
                sheet['A1'].fill = fill_main_transparent
                sheet['B1'].fill = fill_main_transparent
                sheet['A1'].font = font_main_transparent
                sheet['B1'].font = font_main_transparent
                sheet['A2'].border = border_noborder
                sheet['B2'].border = border_noborder
                sheet["B2"].font = font_header
            def format_sheet_phi_coefficients(sheet):
                alignment_left_top = Alignment(indent=0,horizontal='left',vertical='top')
                alignment_left_bottom = Alignment(indent=0,horizontal='left',vertical='bottom')
                fill_strong = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
                fill_light = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                for row in sheet.iter_rows(min_col=1, max_col=1, min_row=1, max_row=sheet.max_row):
                    for cell in row:
                        cell.font = Font(bold=False)
                        cell.border = Border()
                        cell.alignment = alignment_left_top
                for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
                    for cell in col:
                        cell.font = Font(bold=True)
                        cell.border = Border()
                        cell.alignment = alignment_left_bottom
                for col in sheet.columns:
                    col_letter = col[0].column_letter  # Get the column letter (e.g., 'A', 'B', etc.)
                    # sheet.column_dimensions[col_letter].width = 80
                min_row = 2
                min_col = 2
                max_row = sheet.max_row
                max_col = sheet.max_column
                cell_range = f"{sheet.cell(row=min_row, column=min_col).coordinate}:{sheet.cell(row=max_row, column=max_col).coordinate}"
                rule_strong = CellIsRule(operator="greaterThan", formula=["0.5"], fill=fill_strong)
                rule_moderate = CellIsRule(operator="greaterThan", formula=["0.3"], fill=fill_light)
                sheet.conditional_formatting.add(cell_range, rule_strong)
                sheet.conditional_formatting.add(cell_range, rule_moderate)
                sheet.column_dimensions['A'].width = 25
                sheet.row_dimensions[1].height = 21
            def format_sheet_strong_results(sheet):
                alignment_left_top = Alignment(indent=0,horizontal='left',vertical='top')
                alignment_left_bottom = Alignment(indent=0,horizontal='left',vertical='bottom')
                for row in sheet.iter_rows(min_col=1, max_col=1, min_row=1, max_row=sheet.max_row):
                    for cell in row:
                        cell.font = Font(bold=False)
                        cell.border = Border()
                        cell.alignment = alignment_left_top
                for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
                    for cell in col:
                        cell.font = Font(bold=True)
                        cell.border = Border()
                        cell.alignment = alignment_left_bottom
                for col in sheet.columns:
                    col_letter = col[0].column_letter  # Get the column letter (e.g., 'A', 'B', etc.)
                    # sheet.column_dimensions[col_letter].width = 80
                sheet.column_dimensions['A'].width = 35
                sheet.row_dimensions[1].height = 21
            # overview sheet
            df_overview = pd.DataFrame([
                ['',sanitize_text_extract_filename(input_filename)],
                ['spss',input_filename],
                ['run datetime','{dt}'.format(dt=time_start)],
                ['filter',group_filter],
                ['variable select pattern',var_pattern],
            ], columns=['name','value']).set_index("name")
            df_overview.to_excel(writer, sheet_name='overview')
            format_sheet_overview(writer.sheets['overview'])
            # phi_coefficients sheet
            phi_matrix.to_excel(writer, sheet_name='phi_coefficients')
            format_sheet_phi_coefficients(writer.sheets['phi_coefficients'])
            # strong_results sheet
            df_results = pd.DataFrame([[f"{i} vs {j}",phi,count] for i,j,phi,count in strong_results], columns=['pair','phi','cases']).set_index("pair")
            df_results.to_excel(writer, sheet_name='strong_results')
            format_sheet_strong_results(writer.sheets['strong_results'])




        time_finish = datetime.now()
        print('{script_name}: finished at {dt} (elapsed {duration})'.format(dt=time_finish,duration=time_finish-time_start,script_name=script_name))

    except Exception as e:
        # the program is designed to be user-friendly
        # that's why we reformat error messages a little bit
        # stack trace is still printed (I even made it longer to 20 steps!)
        # but the error message itself is separated and printed as the last message again

        # for example, I don't write 'print('File Not Found!');exit(1);', I just write 'raise FileNotFoundErro()'
        print('')
        print('Stack trace:')
        print('')
        traceback.print_exception(e,limit=20)
        print('')
        print('')
        print('')
        print('Error:')
        print('')
        print('{e}'.format(e=e))
        print('')
        exit(1)

if __name__ == '__main__':
    main()